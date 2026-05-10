import tkinter as tk
import openpyxl as op
from tkinter import messagebox, ttk
from datetime import datetime
import os

# Create Excel File
file_name = "favorite_people.xlsx"

if not os.path.exists(file_name):
    wb = op.Workbook()
    ws = wb.active
    ws.title = "Favorite People"

    ws.append(["ID", "Last", "First", "Middle", "BirthYear", "Age"])

    wb.save(file_name)

# Main Window
window = tk.Tk()
window.title("Age Calculator")
window.configure(bg="lightgreen")

# Form Title
title = tk.Label(
    window,
    text="Profile Builder",
    font=("Times New Roman", 14, "bold"),
    bg="lightgreen"
)
title.grid(row=0, column=0, columnspan=6)

# Frame
genframe = tk.Frame(window, bg="lightgreen", bd=2, relief="groove")
genframe.grid(row=1, column=0, columnspan=6, padx=10, pady=10)

# First Name Entry
fname_entry = tk.Entry(genframe, font=("Poppins", 12))
fname_entry.grid(row=2, column=1, columnspan=2, padx=(10, 0), pady=(10, 0))

fname_label = tk.Label(
    genframe,
    text="First Name",
    font=("Poppins", 10, "italic"),
    bg="lightgreen"
)
fname_label.grid(row=3, column=1, columnspan=2)

# Middle Name Entry
mname_entry = tk.Entry(genframe, font=("Poppins", 12))
mname_entry.grid(row=2, column=3, columnspan=2, padx=(10, 0), pady=(10, 0))

mname_label = tk.Label(
    genframe,
    text="Middle Name",
    font=("Poppins", 10, "italic"),
    bg="lightgreen"
)
mname_label.grid(row=3, column=3, columnspan=2)

# Last Name Entry
lname_entry = tk.Entry(genframe, font=("Poppins", 12))
lname_entry.grid(row=2, column=5, columnspan=2, padx=(10, 10), pady=(10, 0))

lname_label = tk.Label(
    genframe,
    text="Last Name",
    font=("Poppins", 10, "italic"),
    bg="lightgreen"
)
lname_label.grid(row=3, column=5, columnspan=2)

# Birthyear Entry
birth_entry = tk.Entry(genframe, font=("Poppins", 12))
birth_entry.grid(row=4, column=1, columnspan=2, padx=(10, 0))

birthyear_label = tk.Label(
    genframe,
    text="Birth Year",
    font=("Poppins", 10, "italic"),
    bg="lightgreen"
)
birthyear_label.grid(row=5, column=2, columnspan=2)

# Function Submit
def submit_data():

    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    birth = birth_entry.get()

    if first == "" or last == "" or birth == "":
        messagebox.showerror("Error", "Please fill all fields.")
        return

    try:
        birth_year = int(birth)
    except:
        messagebox.showerror("Error", "Birth Year must be a number.")
        return

    # Compute Age
    current_year = datetime.now().year
    age = current_year - birth_year

    # Open Excel File
    wb = op.load_workbook(file_name)
    ws = wb.active

    # Auto ID
    new_id = ws.max_row

    # Save Data
    ws.append([new_id, last, first, middle, birth_year, age])

    wb.save(file_name)

    # Display in Treeview
    tree.insert(
        "",
        tk.END,
        values=(new_id, last, first, middle, birth_year, age)
    )

    messagebox.showinfo("Success", "Record Saved Successfully!")

    # Clear Entries
    fname_entry.delete(0, tk.END)
    mname_entry.delete(0, tk.END)
    lname_entry.delete(0, tk.END)
    birth_entry.delete(0, tk.END)

# Update Button
update_btn = tk.Button(window, text="Update")
update_btn.grid(row=6, column=2)

# Submit Button
button = tk.Button(
    window,
    text="Submit",
    font=("Poppins", 12, "bold"),
    bg="lightpink",
    command=submit_data
)
button.grid(row=6, column=0, columnspan=6, pady=(10, 20))

# Delete Button
delete_btn = tk.Button(window, text="Delete", bg="red", fg="white")
delete_btn.grid(row=6, column=3)

# Treeview
tree = ttk.Treeview(
    window,
    columns=("ID", "Last", "First", "Middle", "BirthYear", "Age"),
    show="headings"
)

for col in ("ID", "Last", "First", "Middle", "BirthYear", "Age"):
    tree.heading(col, text=col)

tree.grid(row=7, column=0, columnspan=4)

# Load Existing Data
wb = op.load_workbook(file_name)
ws = wb.active

for row in ws.iter_rows(min_row=2, values_only=True):
    tree.insert("", tk.END, values=row)

window.mainloop()

