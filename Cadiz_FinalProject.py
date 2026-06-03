import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os


# DATABASE SETUP

FILE_NAME = "fitnessDB.xlsx"

if not os.path.exists(FILE_NAME):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fitness"
    ws.append(["ID", "Name", "Weight", "Height", "BMI"])
    wb.save(FILE_NAME)


# FUNCTIONS

def generate_id():
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    return ws.max_row

def calculate_bmi(weight, height):
    height_m = float(height) / 100
    bmi = float(weight) / (height_m ** 2)
    return round(bmi, 2)

def clear_fields():
    id_var.set("")
    name_var.set("")
    weight_var.set("")
    height_var.set("")

def load_data():
    for row in tree.get_children():
        tree.delete(row)

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        tree.insert("", tk.END, values=row)

def add_record():
    name = name_var.get()
    weight = weight_var.get()
    height = height_var.get()

    if not name or not weight or not height:
        messagebox.showerror("Error", "Please fill all fields!")
        return

    try:
        bmi = calculate_bmi(weight, height)
    except:
        messagebox.showerror("Error", "Weight and Height must be numbers!")
        return

    record_id = generate_id()

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    ws.append([record_id, name, float(weight), float(height), bmi])

    wb.save(FILE_NAME)

    messagebox.showinfo("Success", "Record Added Successfully!")
    clear_fields()
    load_data()

def update_record():
    selected = tree.focus()

    if not selected:
        messagebox.showerror("Error", "Select a record first!")
        return

    values = tree.item(selected, "values")
    record_id = int(values[0])

    try:
        bmi = calculate_bmi(weight_var.get(), height_var.get())
    except:
        messagebox.showerror("Error", "Weight and Height must be numbers!")
        return

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if row[0].value == record_id:
            row[1].value = name_var.get()
            row[2].value = float(weight_var.get())
            row[3].value = float(height_var.get())
            row[4].value = bmi
            break

    wb.save(FILE_NAME)

    messagebox.showinfo("Success", "Record Updated Successfully!")
    clear_fields()
    load_data()

def delete_record():
    selected = tree.focus()

    if not selected:
        messagebox.showerror("Error", "Select a record first!")
        return

    confirm = messagebox.askyesno(
        "Confirm Delete",
        "Are you sure you want to delete this record?"
    )

    if confirm:
        values = tree.item(selected, "values")
        record_id = int(values[0])

        wb = load_workbook(FILE_NAME)
        ws = wb.active

        for row_num in range(2, ws.max_row + 1):
            if ws.cell(row=row_num, column=1).value == record_id:
                ws.delete_rows(row_num)
                break

        wb.save(FILE_NAME)

        messagebox.showinfo("Success", "Record Deleted Successfully!")
        clear_fields()
        load_data()

def select_record(event):
    selected = tree.focus()

    if selected:
        values = tree.item(selected, "values")

        id_var.set(values[0])
        name_var.set(values[1])
        weight_var.set(values[2])
        height_var.set(values[3])


# GUI

window = tk.Tk()
window.title("Fitness Progress Tracker")
window.geometry("900x600")
window.configure(bg="#E8F5E9")

title_font = ("Arial", 18, "bold")
label_font = ("Arial", 11, "bold")
button_font = ("Arial", 10, "bold")

id_var = tk.StringVar()
name_var = tk.StringVar()
weight_var = tk.StringVar()
height_var = tk.StringVar()

title_label = tk.Label(
    window,
    text="FITNESS PROGRESS TRACKER",
    font=title_font,
    bg="#E8F5E9",
    fg="#1B5E20"
)
title_label.pack(pady=10)

form_frame = tk.Frame(window, bg="#E8F5E9")
form_frame.pack(pady=10)

tk.Label(form_frame, text="ID", font=label_font, bg="#E8F5E9").grid(row=0, column=0, padx=5, pady=5)
tk.Entry(form_frame, textvariable=id_var, state="readonly").grid(row=0, column=1)

tk.Label(form_frame, text="Name", font=label_font, bg="#E8F5E9").grid(row=1, column=0, padx=5, pady=5)
tk.Entry(form_frame, textvariable=name_var).grid(row=1, column=1)

tk.Label(form_frame, text="Weight (kg)", font=label_font, bg="#E8F5E9").grid(row=2, column=0, padx=5, pady=5)
tk.Entry(form_frame, textvariable=weight_var).grid(row=2, column=1)

tk.Label(form_frame, text="Height (cm)", font=label_font, bg="#E8F5E9").grid(row=3, column=0, padx=5, pady=5)
tk.Entry(form_frame, textvariable=height_var).grid(row=3, column=1)

btn_frame = tk.Frame(window, bg="#E8F5E9")
btn_frame.pack(pady=10)

tk.Button(btn_frame, text="Add Record", width=15, font=button_font,
          bg="#508A79", fg="white",
          command=add_record).grid(row=0, column=0, padx=5)

tk.Button(btn_frame, text="Update Record", width=15, font=button_font,
          bg="#274E6E", fg="white",
          command=update_record).grid(row=0, column=1, padx=5)

tk.Button(btn_frame, text="Delete Record", width=15, font=button_font,
          bg="#EF81BA", fg="white",
          command=delete_record).grid(row=0, column=2, padx=5)

tk.Button(btn_frame, text="Clear", width=15, font=button_font,
          bg="#B753B0", fg="white",
          command=clear_fields).grid(row=0, column=3, padx=5)

style = ttk.Style()
style.theme_use("clam")

style.configure(
    "Treeview",
    font=("Arial", 10),
    rowheight=25
)

style.configure(
    "Treeview.Heading",
    font=("Arial", 11, "bold")
)

columns = ("ID", "Name", "Weight", "Height", "BMI")

tree = ttk.Treeview(window, columns=columns, show="headings", height=15)

for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=150)

tree.pack(fill="both", expand=True, padx=10, pady=10)

tree.bind("<<TreeviewSelect>>", select_record)

load_data()

window.mainloop()
